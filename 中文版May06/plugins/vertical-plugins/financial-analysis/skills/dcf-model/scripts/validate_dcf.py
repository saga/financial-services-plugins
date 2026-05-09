#!/usr/bin/env python3
"""
DCF 模型验证脚本
验证 Excel DCF 模型的公式错误和常见 DCF 错误
"""

import sys
import json
from pathlib import Path
from typing import Optional


class DCFModelValidator:
    """验证 DCF 模型的错误和质量问题"""

    def __init__(self, excel_path: str):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("未安装 openpyxl。运行：pip install openpyxl")

        self.excel_path = excel_path
        self.openpyxl = openpyxl

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"文件未找到：{excel_path}")

        # 加载公式版本和数值版本的工作簿
        self.workbook_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        self.workbook_values = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors = []
        self.warnings = []
        self.info = []
        
    def validate_all(self) -> dict:
        """
        运行所有验证检查

        返回：
            包含验证结果的字典
        """
        from datetime import datetime

        self.check_sheet_structure()
        self.check_formula_errors()
        self.check_dcf_logic()

        results = {
            'file': self.excel_path,
            'validation_date': datetime.now().isoformat(),
            'status': 'PASS' if len(self.errors) == 0 else 'FAIL',
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': self.errors,
            'warnings': self.warnings,
            'info': self.info
        }

        return results
    
    def check_sheet_structure(self):
        """验证必需的工作表是否存在"""
        required_sheets = ['DCF', 'WACC', 'Sensitivity']
        sheet_names = self.workbook_values.sheetnames

        for sheet in required_sheets:
            if sheet not in sheet_names:
                self.warnings.append(f"建议的工作表缺失：{sheet}")
            else:
                self.info.append(f"找到工作表：{sheet}")

    def check_formula_errors(self):
        """检查所有工作表中的 Excel 公式错误"""
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        total_formulas = 0

        for sheet_name in self.workbook_values.sheetnames:
            ws_values = self.workbook_values[sheet_name]
            ws_formulas = self.workbook_formulas[sheet_name]

            for row in ws_values.iter_rows():
                for cell in row:
                    formula_cell = ws_formulas[cell.coordinate]

                    # 统计公式数量
                    if formula_cell.value and isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                        total_formulas += 1

                    # 检查错误值
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                self.errors.append(f"{err} 位于 {location}")
                                break

        # 添加汇总信息
        self.info.append(f"公式总计：{total_formulas}")
        if total_errors == 0:
            self.info.append("✓ 未发现公式错误")
        else:
            self.errors.append(f"公式错误总计：{total_errors}")

        return error_details, total_errors
    
    def check_dcf_logic(self):
        """验证 DCF 特有的逻辑和计算"""
        self._check_terminal_growth_vs_wacc()
        self._check_wacc_range()
        self._check_terminal_value_proportion()

    def _check_terminal_growth_vs_wacc(self):
        """关键检查：终值增长率必须小于 WACC"""
        try:
            dcf_sheet = self.workbook_values['DCF']

            terminal_growth = None
            wacc = None

            # 搜索终值增长率和 WACC 值
            for row in dcf_sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_str = cell.value.lower()
                        if 'terminal' in cell_str and 'growth' in cell_str:
                            # 在相邻单元格中查找值
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    terminal_growth = adjacent
                                    break
                        if 'wacc' in cell_str and wacc is None:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    wacc = adjacent
                                    break

            if terminal_growth is not None and wacc is not None:
                if terminal_growth >= wacc:
                    self.errors.append(
                        f"严重错误：终值增长率 ({terminal_growth:.2%}) >= WACC ({wacc:.2%})。"
                        "这将产生无限价值，数学上不成立。"
                    )
                else:
                    self.info.append(
                        f"✓ 终值增长率 ({terminal_growth:.2%}) < WACC ({wacc:.2%})"
                    )
            else:
                self.warnings.append("无法定位终值增长率和 WACC 值")

        except KeyError:
            self.warnings.append("未找到 DCF 工作表")
        except Exception as e:
            self.warnings.append(f"无法验证终值增长率与 WACC：{str(e)}")

    def _check_wacc_range(self):
        """检查 WACC 是否在合理范围内"""
        try:
            wacc_sheet = self.workbook_values.get('WACC') or self.workbook_values['DCF']
            wacc = None

            for row in wacc_sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if 'wacc' in cell.value.lower():
                            for offset in range(1, 5):
                                adjacent = wacc_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    wacc = adjacent
                                    break

            if wacc is not None:
                if wacc < 0.05 or wacc > 0.20:
                    self.warnings.append(
                        f"WACC ({wacc:.2%}) 超出典型范围（5%-20%）。请验证计算。"
                    )
                else:
                    self.info.append(f"✓ WACC ({wacc:.2%}) 在合理范围内")
            else:
                self.warnings.append("无法定位 WACC 值")

        except Exception as e:
            self.warnings.append(f"无法验证 WACC 范围：{str(e)}")

    def _check_terminal_value_proportion(self):
        """检查终值占企业价值的比例是否合理"""
        try:
            dcf_sheet = self.workbook_values['DCF']

            terminal_value = None
            enterprise_value = None

            for row in dcf_sheet.iter_rows(max_row=200, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_str = cell.value.lower()
                        if 'terminal' in cell_str and 'value' in cell_str and 'pv' in cell_str:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and adjacent > 0:
                                    terminal_value = adjacent
                                    break
                        if 'enterprise' in cell_str and 'value' in cell_str:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and adjacent > 0:
                                    enterprise_value = adjacent
                                    break

            if terminal_value is not None and enterprise_value is not None and enterprise_value > 0:
                proportion = terminal_value / enterprise_value
                if proportion > 0.80:
                    self.warnings.append(
                        f"终值占 EV 的 {proportion:.1%}（通常应在 50-70%）。"
                        "模型可能过度依赖终值假设。"
                    )
                elif proportion < 0.40:
                    self.warnings.append(
                        f"终值占 EV 的 {proportion:.1%}（通常应在 50-70%）。"
                        "检查终值假设是否过于保守。"
                    )
                else:
                    self.info.append(f"✓ 终值占 EV 的 {proportion:.1%}")
            else:
                self.warnings.append("无法定位终值和企业价值")

        except Exception as e:
            self.warnings.append(f"无法验证终值占比：{str(e)}")
    


def validate_dcf_model(excel_path: str) -> dict:
    """
    验证 DCF 模型 Excel 文件

    参数：
        excel_path: Excel DCF 模型路径

    返回：
        包含验证结果的字典
    """
    validator = DCFModelValidator(excel_path)
    return validator.validate_all()


def main():
    """命令行接口"""
    if len(sys.argv) < 2:
        print("用法：python validate_dcf.py <excel文件> [output.json]")
        print("\n验证 DCF 模型：")
        print("  - 公式错误（#REF!、#DIV/0! 等）")
        print("  - 终值增长率 < WACC（关键检查）")
        print("  - WACC 在合理范围内（5-20%）")
        print("  - 终值占 EV 比例（40-80%）")
        print("\n返回含错误、警告和信息的 JSON")
        print("\n示例：python validate_dcf.py model.xlsx")
        print("示例：python validate_dcf.py model.xlsx results.json")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        results = validate_dcf_model(excel_file)

        # 打印结果
        print(json.dumps(results, indent=2))

        # 如指定则保存到文件
        if output_file:
            with open(output_file, 'w') as f:
                json.dump(results, f, indent=2)

        # 验证失败时返回错误码
        sys.exit(0 if results['status'] == 'PASS' else 1)

    except Exception as e:
        error_result = {
            'file': excel_file,
            'status': 'ERROR',
            'error': str(e)
        }
        print(json.dumps(error_result, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()